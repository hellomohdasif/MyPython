# importing openpyxl module
import openpyxl
import whois
from xlrd import open_workbook
from xlutils.copy import copy
# Give the location of the file
path = "C:\\Users\\Asif\Dropbox\\data.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    domain=(cell_obj.value)
    w = whois.whois(domain)
    wx=(w.expiration_date)
    wxx=str(wx)
    sub_str = " "

    # slicing off after length computation
    finalex = wxx[:wxx.index(sub_str) + len(sub_str)]
    print(finalex)


    rb = open_workbook(r"E:\main\all.xls")
    wb = copy(rb)

    s = wb.get_sheet(9)
    s.write(i-i, 0,domain)


    s.write(i-i, 1, finalex)


    wb.save('C:\\Users\\Asif\\Desktop\\Main\\all.xls')



